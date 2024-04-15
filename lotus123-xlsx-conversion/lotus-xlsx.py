import xlrd
from openpyxl import Workbook
import os

def transform_file_path(file_path):

    components = file_path.split(os.sep)
    
    preservation_index = components.index('preservation')
    
    components[preservation_index] = r"access/nearline"
    access = r"access/nearline"
    
    new_file_path = os.sep.join(components[:-1])  # Remove the last directory
    #new_file_path = os.path.join(new_file_path, access)
    print(new_file_path)
    return new_file_path

def find_subfolder(dup):

    if not os.path.exists(start_path):
        print(f"Path {start_path} does not exist.")
        return []

    contents = os.listdir(dup)

    for content in contents:
        
        content_path = os.path.join(dup, content)
        
        if os.path.isdir(content_path):
            print(f"Entering subfolder: {content_path}")
            create = content_path.replace("preservation", "access/nearline")
            os.makedirs(create, exist_ok=True)
            find_subfolder(content_path)

        else:
            print(os.path.dirname(content_path))


def convert_files(folder_path):
    # Check if the specified folder exists
    if not os.path.exists(folder_path):
        print(f"Folder {folder_path} does not exist.")
        return
    # Get the list of subfolders in the specified folder
    # subfolders = [f.path for f in os.scandir(folder_path) if f.is_dir()]
    root_contents = os.listdir(folder_path)
    i = 0
    for content in root_contents:
        content_path = os.path.join(folder_path, content)
        print(content_path)        
        i = i+1
    if(content.endswith('.wk1') or content.endswith('.WK1')):
        try: 
                
            input_file = content_path
            name = content.replace('.wk1', '')
            name = name.replace('.WK1', '')
            output_file_name = f'{name}.xlsx'
            access_path = transform_file_path(folder_path)
            print(access_path)
            output_file = os.path.join(access_path, output_file_name)
            print(output_file)
            convert_to_xlsx(input_file, output_file)
    
        except Exception as E:
            print(E)
            print("Could not convert!")

def convert_to_xlsx(input_file, output_file):
    # Open Lotus 1-2-3 file
    wb_lotus = xlrd.open_workbook(input_file)

    # Create a new workbook
    wb_xlsx = Workbook()

    # Iterate over all sheets in Lotus 1-2-3 file
    for lotus_sheet_index in range(wb_lotus.nsheets):
        lotus_sheet = wb_lotus.sheet_by_index(lotus_sheet_index)
        ws_xlsx = wb_xlsx.create_sheet(title=lotus_sheet.name)

        # Copy data from Lotus 1-2-3 to .xlsx
        for row_index in range(lotus_sheet.nrows):
            for col_index in range(lotus_sheet.ncols):
                cell_value = lotus_sheet.cell_value(row_index, col_index)
                ws_xlsx.cell(row=row_index + 1, column=col_index + 1).value = cell_value

    # Remove default sheet created by openpyxl
    wb_xlsx.remove(wb_xlsx.worksheets[0])

    # Save .xlsx file
    wb_xlsx.save(output_file)

start_path = r"Z:/Shri - Script Test Folders/Emiquon/4308001copy_UNPROCESSED/preservation"  # Change this to the path of the folder containing lotus files, SHOULD BE A PRESERVATION FOLDER
#xlsx_file = r"C:/Users/pal10/Desktop/DISKS/LHB1.xlsx"
find_subfolder(start_path)
convert_files(start_path)